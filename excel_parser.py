import json
import logging
import re
import time
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from ai_detector import AIStructureDetector
from groq import Groq
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Constants
API_REQUEST_TIMEOUT = 30  # seconds
API_MAX_RETRIES = 3
API_RETRY_DELAY = 1  # seconds (base for exponential backoff)


class AIEnhancedParser:
    """Lightweight AI helper for extracting credentials from Notes fields."""

    def __init__(self, api_key: str, model: str):
        self.client = Groq(api_key=api_key)
        self.model = model

    def extract_credentials(
        self, items: List[Dict[str, str]]
    ) -> Dict[int, Dict[str, str]]:
        """
        Extract credentials from notes.

        Args:
            items: [{"index": int, "title": str, "notes": str}, ...]

        Returns:
            {index: {"username": str, "password": str, "notes_clean": str}}
        """
        if not items:
            return {}

        # Limit size to keep prompt small
        items = items[:30]

        prompt = (
            "You extract credentials from Notes fields for Passbolt import.\n"
            "Input is a JSON array of items with index, title, and notes.\n"
            "Return ONLY a JSON object in this format:\n"
            '{ "records": [ { "index": 0, "username": "...", "password": "...", "notes_clean": "..." } ] }\n\n'
            "Rules:\n"
            "- Only include records where you are confident credentials exist.\n"
            "- Do NOT invent data.\n"
            "- Username and password must be exact substrings from notes.\n"
            "- notes_clean should be notes with extracted credentials removed (optional).\n"
            '- If none found, return {"records": []}.\n\n'
            f"ITEMS:\n{json.dumps(items, ensure_ascii=False)}"
        )

        # Call Groq API with retry logic
        response = None
        last_error = None

        for attempt in range(API_MAX_RETRIES):
            try:
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                    response_format={"type": "json_object"},
                    temperature=0.1,
                    max_tokens=800,
                    timeout=API_REQUEST_TIMEOUT,
                )
                break  # Success
            except Exception as e:
                last_error = e
                logger.warning(
                    f"API request failed (attempt {attempt + 1}/{API_MAX_RETRIES}): {e}"
                )
                if attempt < API_MAX_RETRIES - 1:
                    # Exponential backoff
                    delay = API_RETRY_DELAY * (2**attempt)
                    logger.info(f"Retrying in {delay} seconds...")
                    time.sleep(delay)

        if response is None:
            logger.error(
                f"Failed to call Groq API after {API_MAX_RETRIES} attempts. "
                f"Last error: {last_error}"
            )
            return {}  # Return empty result instead of raising

        text = response.choices[0].message.content.strip()

        # Extract JSON object if wrapped
        json_match = re.search(r"```(?:json)?\s*({.*?})\s*```", text, re.DOTALL)
        if json_match:
            text = json_match.group(1)

        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            # Attempt minor fixes
            fixed = text.replace("'", '"')
            fixed = re.sub(r",\s*}", "}", fixed)
            fixed = re.sub(r",\s*]", "]", fixed)
            try:
                data = json.loads(fixed)
            except Exception:
                return {}

        if not isinstance(data, dict):
            return {}

        records = data.get("records", [])
        if not isinstance(records, list):
            return {}

        updates: Dict[int, Dict[str, str]] = {}
        for rec in records:
            if not isinstance(rec, dict):
                continue
            try:
                idx = int(rec.get("index"))
            except Exception:
                continue
            username = str(rec.get("username", "") or "").strip()
            password = str(rec.get("password", "") or "").strip()
            notes_clean = rec.get("notes_clean")
            if notes_clean is not None:
                notes_clean = str(notes_clean)
            if username or password or notes_clean is not None:
                updates[idx] = {
                    "username": username,
                    "password": password,
                    "notes_clean": notes_clean,
                }

        return updates


class ExcelParser:
    """Parser for complex Excel files with support for merged cells, multi-row headers, etc."""

    PASSBOLT_COLUMNS = ["Group", "Title", "Username", "Password", "URL", "Notes"]

    # Keywords to identify column types
    COLUMN_KEYWORDS = {
        "Title": [
            "name",
            "title",
            "resource",
            "service",
            "app",
            "application",
            "system",
            "server",
            "назва",
            "ім'я",
        ],
        "Username": [
            "user",
            "login",
            "username",
            "account",
            "email",
            "користувач",
            "логін",
        ],
        "Password": ["pass", "password", "pwd", "secret", "пароль"],
        "URL": [
            "url",
            "link",
            "address",
            "host",
            "ip",
            "domain",
            "website",
            "адреса",
            "хост",
        ],
        "Group": [
            "group",
            "category",
            "folder",
            "space",
            "env",
            "environment",
            "dept",
            "department",
            "група",
            "категорія",
        ],
        "Notes": [
            "notes",
            "comment",
            "description",
            "info",
            "extra",
            "additional",
            "примітки",
            "опис",
        ],
    }

    def __init__(
        self,
        file_path: str,
        groq_api_key: Optional[str] = None,
        model: str = "llama-3.3-70b-versatile",
    ):
        self.file_path = file_path
        self.wb = None
        self.df = None
        self.raw_data = []
        self.groq_api_key = groq_api_key
        self.model = model
        self.ai_detector = (
            AIStructureDetector(api_key=groq_api_key, model=model)
            if groq_api_key
            else None
        )

    def load_excel(self, sheet_name: Optional[str] = None) -> None:
        """Load Excel file with support for complex structures."""
        self.wb = load_workbook(self.file_path, data_only=True)

        if sheet_name:
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.active

        self._parse_worksheet(ws)

    def _parse_worksheet(self, ws) -> None:
        """Parse worksheet handling merged cells and complex structures."""
        # Handle merged cells
        merged_cells = {}
        for merged_range in ws.merged_cells.ranges:
            value = ws[merged_range.min_row][merged_range.min_col - 1].value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells[(row, col)] = value

        # Read all data with merged cell handling
        data = []
        max_row = ws.max_row
        max_col = ws.max_column

        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value

                # Check merged cells
                if cell_value is None:
                    cell_value = merged_cells.get((row_idx, col_idx))

                row_data.append(cell_value)
            data.append(row_data)

        # Remove completely empty rows
        data = [row for row in data if any(cell is not None for cell in row)]

        if not data:
            raise ValueError("Empty worksheet")

        self.raw_data = data
        self.df = pd.DataFrame(data)

    def detect_header_row(self) -> int:
        """Detect which row contains headers (0-indexed)."""
        for row_idx, row in enumerate(self.raw_data):
            # Count how many cells in this row match header keywords
            header_matches = 0
            total_non_empty = 0

            for cell in row:
                if cell and isinstance(cell, str):
                    total_non_empty += 1
                    cell_lower = cell.lower()
                    for col_type, keywords in self.COLUMN_KEYWORDS.items():
                        if any(kw in cell_lower for kw in keywords):
                            header_matches += 1
                            break

            # A header row should have multiple column matches
            # and at least 50% of non-empty cells should match keywords
            if total_non_empty >= 2 and header_matches >= 2:
                return row_idx
            # Also check if row has 3+ header keyword matches
            if header_matches >= 3:
                return row_idx

        return 0  # Default to first row if no headers detected

    def identify_columns(self, header_row_idx: int) -> Dict[str, int]:
        """Map Passbolt columns to Excel column indices."""
        column_mapping = {}
        headers = self.raw_data[header_row_idx]

        for col_idx, header in enumerate(headers):
            if header is None:
                continue

            header_lower = str(header).lower()

            for passbolt_col, keywords in self.COLUMN_KEYWORDS.items():
                for keyword in keywords:
                    if keyword in header_lower:
                        if passbolt_col not in column_mapping:
                            column_mapping[passbolt_col] = col_idx
                            break

        # If no columns mapped, try positional mapping (common patterns)
        if not column_mapping:
            column_mapping = self._positional_mapping(len(headers))

        return column_mapping

    def _positional_mapping(self, num_cols: int) -> Dict[str, int]:
        """Fallback positional mapping for sheets without headers.

        Common pattern: [Name, IP/URL, Location, Credentials/Notes, ...]
        """
        column_mapping = {}

        # Col 0: Title/Name
        if num_cols >= 1:
            column_mapping["Title"] = 0

        # Col 1: IP/URL
        if num_cols >= 2:
            column_mapping["URL"] = 1

        # Col 2: Location/Notes
        if num_cols >= 3:
            column_mapping["Notes"] = 2

        # Col 3: Username/Password
        if num_cols >= 4:
            column_mapping["Username"] = 3

        return column_mapping

    def _is_group_header_row(self, row: List) -> Optional[str]:
        """
        Check if row is a group/section header (metadata row).

        Group headers typically:
        - Have only first column filled
        - Rest columns are empty or None
        - First column contains group name like "для Бобкова", "Кластре ІТ", "Demo-k8s"

        Returns:
            Group name if it's a group header, None otherwise
        """
        if not row or len(row) < 2:
            return None

        first_col = row[0]

        # First column should have a value
        if not first_col or not isinstance(first_col, str):
            return None

        first_col = str(first_col).strip()

        # Should not be empty or look like a server name with IP
        if not first_col or first_col.startswith("http"):
            return None

        # Check if rest of columns are empty
        rest_empty = all(
            not cell or (isinstance(cell, str) and not cell.strip()) for cell in row[1:]
        )

        if rest_empty:
            # This is likely a group header
            return first_col

        # Also check for common group header patterns
        group_keywords = ["для ", "кластер", "cluster", "group", "категорія"]
        first_col_lower = first_col.lower()

        for keyword in group_keywords:
            if keyword in first_col_lower:
                return first_col

        return None

    def _is_server_row(self, row: List) -> bool:
        """Check if row looks like a server entry (name + IP pattern)."""
        if len(row) < 2:
            return False

        # Guard: header-like rows are not servers
        header_matches = 0
        for cell in row:
            if cell and isinstance(cell, str):
                cell_lower = cell.lower()
                for _, keywords in self.COLUMN_KEYWORDS.items():
                    if any(kw in cell_lower for kw in keywords):
                        header_matches += 1
                        break
        if header_matches >= 2:
            return False

        name = row[0]
        ip_or_url = row[1] if len(row) > 1 else None

        # First cell should be a name (not empty, not URL)
        if not name or not isinstance(name, str):
            return False

        name = str(name).strip()
        if not name or name.startswith("http"):
            return False

        # If second column looks like a password (not IP, not empty), it's credentials not a server
        if ip_or_url:
            ip_str = str(ip_or_url).strip()
            # If it's an IP, it's a server row
            if re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", ip_str):
                return True
            # If it looks like a password (special chars, short), it's credentials
            if len(name) <= 15 and len(ip_str) >= 4 and not ip_str.startswith("http"):
                return False  # This is credentials, not a server

        # No IP in second column - could be a server without IP or credentials
        # Check if it looks like credentials (short name + password-like value)
        if ip_or_url and len(name) <= 15 and len(str(ip_or_url).strip()) >= 4:
            return False

        return bool(name)

    def _is_credentials_row(self, row: List) -> Optional[Tuple[str, str]]:
        """Check if row contains credentials (username/password)."""
        if len(row) < 2:
            return None

        # Guard: header-like rows are not credentials
        header_matches = 0
        for cell in row:
            if cell and isinstance(cell, str):
                cell_lower = cell.lower()
                for _, keywords in self.COLUMN_KEYWORDS.items():
                    if any(kw in cell_lower for kw in keywords):
                        header_matches += 1
                        break
        if header_matches >= 2:
            return None

        col0 = row[0]
        col1 = row[1]

        if not col0 or not col1:
            return None

        col0_str = str(col0).strip()
        col1_str = str(col1).strip()

        # Skip if looks like URL
        if col0_str.startswith("http") or col1_str.startswith("http"):
            return None

        # Skip if looks like IP
        if re.match(r"^\d{1,3}\.", col0_str) or re.match(r"^\d{1,3}\.", col1_str):
            return None

        # col0 should be a short username (not a server name)
        # col1 should look like a password (has special chars, mixed case, etc.)
        if len(col0_str) <= 15 and len(col1_str) >= 4:
            return (col0_str, col1_str)

        return None

    def _is_url_row(self, row: List) -> Optional[str]:
        """Check if row contains a URL."""
        if not row:
            return None

        for cell in row:
            if cell and isinstance(cell, str):
                cell_str = str(cell).strip()
                if cell_str.startswith("http://") or cell_str.startswith("https://"):
                    return cell_str
                # Also check for URL-like patterns
                if re.match(r"^https?://", cell_str, re.IGNORECASE):
                    return cell_str

        return None

    def parse_vertical_format(self) -> List[Dict[str, str]]:
        """
        Parse vertical format where:
        - Server rows: [Name, IP, Port/Notes, ...]
        - Credential rows: [Username, Password, ...]
        - URL rows: [URL, ...]
        - Group headers: [Group Name, None, None, ...]

        Credentials and URLs apply to the last server row above them.
        Group headers apply to all servers below them until next group header.
        Order can be: Group -> Server -> Creds -> URL -> Creds -> URL
        All creds and URLs between servers apply to the previous server.
        """
        results = []
        current_server = None
        pending_credentials = []
        pending_url = None
        current_group = "Imported"  # Default group

        def finalize_server(server, creds, url):
            """Apply pending creds and URL to server."""
            if not server:
                return

            # Apply credentials
            if creds:
                if len(creds) == 1:
                    server["Username"] = creds[0][0]
                    server["Password"] = creds[0][1]
                else:
                    # Multiple credentials - use last as main, others in notes
                    all_creds_str = " | ".join([f"{u}: {p}" for u, p in creds])
                    if server["Notes"]:
                        server["Notes"] += f" | {all_creds_str}"
                    else:
                        server["Notes"] = all_creds_str

                    # Set last credential as primary
                    server["Username"] = creds[-1][0]
                    server["Password"] = creds[-1][1]

            # Apply URL
            if url:
                server["URL"] = url

            return server

        for row_idx, row in enumerate(self.raw_data):
            # Clean row - remove None values for easier checking
            clean_row = [c for c in row if c is not None and str(c).strip()]
            if not clean_row:
                continue

            # Check for group header row FIRST
            group_name = self._is_group_header_row(row)
            if group_name:
                current_group = group_name
                continue

            # Check for URL row
            url = self._is_url_row(row)
            if url:
                pending_url = url
                continue

            # Check for credentials row
            creds = self._is_credentials_row(row)
            if creds:
                pending_credentials.append(creds)
                continue

            # Check for server row
            if self._is_server_row(row):
                # Save previous server if exists
                if current_server:
                    current_server = finalize_server(
                        current_server, pending_credentials, pending_url
                    )
                    results.append(current_server)

                # Create new server entry
                name = str(row[0]).strip()
                ip = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                notes = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                current_server = {
                    "Group": current_group,
                    "Title": name,
                    "Username": "",
                    "Password": "",
                    "URL": self.normalize_url(ip) if ip else "",
                    "Notes": notes,
                }

                # Reset pending items for new server
                pending_url = None
                pending_credentials = []

        # Don't forget the last server
        if current_server:
            current_server = finalize_server(
                current_server, pending_credentials, pending_url
            )
            results.append(current_server)

        return results

    def normalize_url(self, url: str) -> str:
        """Normalize URL to include protocol."""
        if not url:
            return ""

        url = str(url).strip()

        # Skip if already has protocol
        if url.startswith(("http://", "https://", "ftp://")):
            return url

        # Check if it's an IP address
        ip_pattern = r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}"
        if re.match(ip_pattern, url):
            return f"http://{url}"

        # Check if it looks like a domain
        if "." in url and not url.startswith("@"):
            return f"https://{url}"

        return url

    def extract_data(
        self, column_mapping: Dict[str, int], start_row: int
    ) -> List[Dict[str, str]]:
        """Extract and map data to Passbolt format."""
        results = []
        current_group = "Imported"  # Default group

        for row_idx in range(start_row, len(self.raw_data)):
            row = self.raw_data[row_idx]

            # Skip completely empty rows
            if not any(cell for cell in row):
                continue

            # Check for group header row (skip it, but remember the group name)
            group_name = self._is_group_header_row(row)
            if group_name:
                current_group = group_name
                continue

            record = {col: "" for col in self.PASSBOLT_COLUMNS}
            record["Group"] = current_group  # Apply current group

            # Map columns
            for passbolt_col, col_idx in column_mapping.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        record[passbolt_col] = str(value).strip()

            # FIRST: Check ALL columns for combined credentials format
            # This must happen BEFORE adding to Notes
            # Skip column 0 (Title) to avoid false positives
            for col_idx, value in enumerate(row):
                if col_idx == 0:  # Skip Title column
                    continue
                if value is not None:
                    val_str = str(value).strip()
                    # Check if it looks like credentials (username/password format)
                    if (
                        not val_str.startswith("http")
                        and not val_str.startswith("192.168.")
                        and not val_str.startswith("172.16.")
                        and not val_str.startswith("10.")
                        and "/" in val_str
                    ):
                        parts = val_str.split("/", 1)
                        if (
                            len(parts) == 2
                            and len(parts[0].strip()) <= 15
                            and len(parts[1].strip()) >= 4
                            and " " not in parts[0]  # No spaces in username
                        ):
                            # Likely credentials - only set if not already set
                            if not record["Username"]:
                                record["Username"] = parts[0].strip()
                                record["Password"] = parts[1].strip()

            # Collect unmapped data into Notes
            extra_data = []
            for col_idx, value in enumerate(row):
                if col_idx not in column_mapping.values() and value is not None:
                    val_str = str(value).strip()
                    # Skip if it looks like credentials (already processed)
                    if "/" in val_str and not val_str.startswith("http"):
                        parts = val_str.split("/", 1)
                        if (
                            len(parts) == 2
                            and len(parts[0].strip()) <= 15
                            and len(parts[1].strip()) >= 4
                            and " " not in parts[0]
                        ):
                            continue  # Skip credentials in Notes
                    extra_data.append(val_str)

            if extra_data:
                existing_notes = record["Notes"]
                new_notes = " | ".join(extra_data)
                record["Notes"] = (
                    f"{existing_notes} | {new_notes}" if existing_notes else new_notes
                )

            # Skip rows without Title
            if not record["Title"]:
                continue

            # Normalize URL
            record["URL"] = self.normalize_url(record["URL"])

            # Remove credentials pattern from Notes if already in Username/Password
            if record["Username"] and record["Password"] and record["Notes"]:
                cred_pattern = f"{record['Username']}/{record['Password']}"
                if cred_pattern in record["Notes"]:
                    record["Notes"] = record["Notes"].replace(cred_pattern, "").strip()
                    # Clean up extra separators
                    record["Notes"] = re.sub(r"\s*\|\s*\|\s*", " | ", record["Notes"])
                    record["Notes"] = re.sub(r"^\s*\|\s*", "", record["Notes"])
                    record["Notes"] = re.sub(r"\s*\|\s*$", "", record["Notes"])

            # Set default Group
            if not record["Group"]:
                record["Group"] = "Imported"

            results.append(record)

        return results

    def parse(
        self, sheet_name: Optional[str] = None, use_ai: bool = True
    ) -> List[Dict[str, str]]:
        """
        Full parse pipeline with optional AI structure detection.

        Args:
            sheet_name: Optional sheet name to parse
            use_ai: Whether to use AI for structure detection (default: True)
        """
        # Avoid reloading if data already present for this sheet
        if not self.raw_data:
            self.load_excel(sheet_name)
        elif sheet_name and self.wb and self.wb.active.title != sheet_name:
            self.load_excel(sheet_name)

        # Try AI-based structure detection first
        if use_ai and self.ai_detector:
            try:
                ai_result = self._detect_structure_with_ai(sheet_name)
                if ai_result:
                    return ai_result
            except Exception as e:
                # Fall back to rule-based detection
                print(f"AI detection failed: {e}, using fallback...")

        # Try vertical format only if both server and credential/url patterns exist
        if len(self.raw_data) >= 2:
            has_server = False
            has_creds_or_url = False
            for row in self.raw_data[:5]:
                if self._is_server_row(row):
                    has_server = True
                if self._is_credentials_row(row) or self._is_url_row(row):
                    has_creds_or_url = True
            has_vertical_pattern = has_server and has_creds_or_url

            if has_vertical_pattern:
                vertical_data = self.parse_vertical_format()
                if vertical_data:
                    return vertical_data

        # Fall back to standard table format
        header_row = self.detect_header_row()
        column_mapping = self.identify_columns(header_row)

        # Check if detected "header row" looks like actual data
        headers = self.raw_data[header_row]
        is_header_actually_data = False

        if "Title" in column_mapping and "URL" in column_mapping:
            title_col = column_mapping["Title"]
            url_col = column_mapping["URL"]

            if title_col < len(headers) and url_col < len(headers):
                title_val = headers[title_col]
                url_val = headers[url_col]

                if url_val and re.match(
                    r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", str(url_val)
                ):
                    is_header_actually_data = True

        if is_header_actually_data:
            column_mapping = self._positional_mapping(len(self.raw_data[0]))
            data = self.extract_data(column_mapping, 0)
        else:
            data = self.extract_data(column_mapping, header_row + 1)

        # AI Enhancement: Detect combined credentials if none found
        with_creds = [r for r in data if r.get("Username")]
        if len(with_creds) == 0 and self.groq_api_key:
            data = self._ai_detect_credentials(data)

        return data

    def _ai_detect_credentials(
        self, data: List[Dict[str, str]]
    ) -> List[Dict[str, str]]:
        """
        Use AI to detect combined credentials in Notes field.

        Sometimes credentials are in format "sa/password123" in the Notes column.
        AI can detect this pattern better than rules.
        """
        # Sample data for AI
        sample_notes = [r.get("Notes", "") for r in data[:20] if r.get("Notes")]

        if not sample_notes:
            return data

        # Check each record for potential credentials in Notes
        for record in data:
            notes = record.get("Notes", "")
            if not notes:
                continue

            # Check for combined credentials pattern
            if "/" in notes and not notes.startswith("http"):
                # Try to extract credentials
                parts = notes.split("/", 1)
                if (
                    len(parts) == 2
                    and len(parts[0].strip()) <= 15
                    and len(parts[1].strip()) >= 4
                    and " " not in parts[0].strip()
                ):
                    # Likely credentials
                    record["Username"] = parts[0].strip()
                    record["Password"] = parts[1].strip()
                    # Remove from notes
                    record["Notes"] = notes.replace(
                        f"{parts[0]}/{parts[1]}", ""
                    ).strip()
                    record["Notes"] = re.sub(r"\s*\|\s*\|\s*", " | ", record["Notes"])
                    record["Notes"] = record["Notes"].strip(" |")

        # If still no credentials found, try AI extraction from Notes
        has_creds = any(r.get("Username") or r.get("Password") for r in data)
        if not has_creds and self.groq_api_key:
            candidates = []
            for idx, rec in enumerate(data):
                notes = (rec.get("Notes") or "").strip()
                if not notes:
                    continue
                notes_lower = notes.lower()
                if (
                    "/" in notes
                    or "login" in notes_lower
                    or "password" in notes_lower
                    or "пароль" in notes_lower
                    or "логін" in notes_lower
                ):
                    candidates.append(
                        {
                            "index": idx,
                            "title": rec.get("Title", ""),
                            "notes": notes,
                        }
                    )
                if len(candidates) >= 30:
                    break

            if candidates:
                try:
                    ai_parser = AIEnhancedParser(self.groq_api_key, self.model)
                    updates = ai_parser.extract_credentials(candidates)
                    for idx, upd in updates.items():
                        if idx < 0 or idx >= len(data):
                            continue
                        rec = data[idx]
                        if upd.get("username") and not rec.get("Username"):
                            rec["Username"] = upd["username"]
                        if upd.get("password") and not rec.get("Password"):
                            rec["Password"] = upd["password"]
                        if upd.get("notes_clean") is not None:
                            rec["Notes"] = upd["notes_clean"]
                except Exception as e:
                    print(f"AI credential extraction failed: {e}")

        return data

    def _detect_structure_with_ai(
        self, sheet_name: Optional[str] = None
    ) -> Optional[List[Dict[str, str]]]:
        """
        Use AI to detect structure and parse accordingly.

        Returns:
            Parsed data if AI detection succeeds, None otherwise
        """
        if not self.ai_detector:
            return None

        # Get sheet name
        if not sheet_name:
            sheet_name = self.wb.active.title if self.wb else "Sheet1"

        # Take sample rows for AI analysis
        sample = self.raw_data[:15]

        # Detect structure
        result = self.ai_detector.analyze(sample, sheet_name)

        # Use AI-detected structure
        header_row = result.get("header_row")
        columns = result.get("columns", {})
        is_vertical = result.get("is_vertical_format", False)

        # Filter out null columns
        column_mapping = {k: v for k, v in columns.items() if v is not None}

        # Handle vertical format
        if is_vertical:
            return self.parse_vertical_format()

        # Handle no headers
        if header_row is None:
            return self.extract_data(column_mapping, 0)

        # Standard format with headers
        return self.extract_data(column_mapping, header_row + 1)

    def get_sheet_names(self) -> List[str]:
        """Get all sheet names."""
        if not self.wb:
            self.wb = load_workbook(self.file_path, data_only=True)
        return self.wb.sheetnames

    def export_to_csv(self, data: List[Dict[str, str]], output_path: str) -> None:
        """Export parsed data to CSV for Passbolt import."""
        import csv

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=self.PASSBOLT_COLUMNS,
                quoting=csv.QUOTE_ALL,
                lineterminator="\r\n",
            )
            writer.writeheader()
            writer.writerows(data)


def parse_excel_to_passbolt(
    input_path: str, output_path: str, sheet_name: Optional[str] = None
) -> int:
    """
    Parse Excel file and export to Passbolt CSV format.

    Args:
        input_path: Path to input Excel file
        output_path: Path for output CSV file
        sheet_name: Optional sheet name to parse (default: first sheet)

    Returns:
        Number of records exported
    """
    parser = ExcelParser(input_path)
    data = parser.parse(sheet_name)
    parser.export_to_csv(data, output_path)
    return len(data)


if __name__ == "__main__":
    # Example usage
    import sys

    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
        sheet = sys.argv[3] if len(sys.argv) > 3 else None

        count = parse_excel_to_passbolt(input_file, output_file, sheet)
        print(f"Exported {count} records to {output_file}")
    else:
        print("Usage: python excel_parser.py <input.xlsx> <output.csv> [sheet_name]")
