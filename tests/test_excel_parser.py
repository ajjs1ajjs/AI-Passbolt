import os
import tempfile
import unittest

from excel_parser import ExcelParser
from openpyxl import Workbook


def _make_xlsx(rows, merges=None):
    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    if merges:
        for rng in merges:
            ws.merge_cells(rng)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


class TestExcelParser(unittest.TestCase):
    def _parse(self, rows, sheet_name=None, use_ai=False):
        path = _make_xlsx(rows)
        try:
            parser = ExcelParser(path)
            data = parser.parse(sheet_name=sheet_name, use_ai=use_ai)
            return data
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass

    def test_standard_header_table(self):
        rows = [
            ["Title", "Username", "Password", "URL", "Notes", "Group"],
            ["Web", "admin", "secret", "192.168.1.10", "Main server", "Prod"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        rec = data[0]
        self.assertEqual(rec["Title"], "Web")
        self.assertEqual(rec["Username"], "admin")
        self.assertEqual(rec["Password"], "secret")
        self.assertEqual(rec["Group"], "Prod")
        self.assertEqual(rec["Notes"], "Main server")
        self.assertEqual(rec["URL"], "http://192.168.1.10")

    def test_group_headers_in_table(self):
        rows = [
            ["Title", "Username", "Password", "URL", "Notes"],
            ["Group A", None, None, None, None],
            ["Srv1", "user1", "pass1", "10.0.0.1", "n1"],
            ["Srv2", "", "", "10.0.0.2", ""],
            ["Group B", None, None, None, None],
            ["Srv3", "user3", "pass3", "10.0.0.3", ""],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 3)
        self.assertEqual(data[0]["Group"], "Group A")
        self.assertEqual(data[1]["Group"], "Group A")
        self.assertEqual(data[2]["Group"], "Group B")

    def test_vertical_format(self):
        rows = [
            ["Group A", None, None],
            ["Server1", "192.168.0.1", "note1"],
            ["admin", "P@ss123", None],
            ["https://srv1.example.com", None, None],
            ["Server2", "192.168.0.2", None],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 2)
        first = data[0]
        second = data[1]
        self.assertEqual(first["Group"], "Group A")
        self.assertEqual(first["Title"], "Server1")
        self.assertEqual(first["Username"], "admin")
        self.assertEqual(first["Password"], "P@ss123")
        self.assertEqual(first["URL"], "https://srv1.example.com")
        self.assertEqual(second["Title"], "Server2")
        self.assertEqual(second["URL"], "http://192.168.0.2")

    def test_combined_credentials_in_notes(self):
        rows = [
            ["Title", "Notes", "URL"],
            ["Srv1", "admin/Pass123", "192.168.1.1"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        rec = data[0]
        self.assertEqual(rec["Username"], "admin")
        self.assertEqual(rec["Password"], "Pass123")
        self.assertEqual(rec["URL"], "http://192.168.1.1")
        self.assertEqual(rec["Notes"], "")

    def test_domain_url_normalization(self):
        rows = [
            ["Title", "URL"],
            ["App", "example.com"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["URL"], "https://example.com")

    def test_empty_file(self):
        rows = []
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 0)

    def test_header_only_file(self):
        rows = [["Title", "Username", "Password", "URL", "Notes", "Group"]]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 0)

    def test_multiple_sheets(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Production"
        ws1.append(["Title", "Username", "Password", "URL"])
        ws1.append(["Server1", "admin", "pass1", "192.168.1.1"])

        ws2 = wb.create_sheet("Development")
        ws2.append(["Title", "Username", "Password", "URL"])
        ws2.append(["Server2", "dev", "pass2", "192.168.1.2"])

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        wb.save(tmp.name)

        try:
            parser = ExcelParser(tmp.name)
            data_prod = parser.parse(sheet_name="Production", use_ai=False)
            data_dev = parser.parse(sheet_name="Development", use_ai=False)

            self.assertEqual(len(data_prod), 1)
            self.assertEqual(data_prod[0]["Title"], "Server1")
            self.assertEqual(len(data_dev), 1)
            self.assertEqual(data_dev[0]["Title"], "Server2")
        finally:
            os.unlink(tmp.name)

    def test_merged_cells(self):
        rows = [
            ["Group", None, None, None],
            ["Server", "User", "Pass", "URL"],
            ["Srv1", "admin", "pass1", "192.168.1.1"],
        ]
        merges = ["A1:D1"]
        path = _make_xlsx(rows, merges)
        try:
            parser = ExcelParser(path)
            data = parser.parse(use_ai=False)
            self.assertEqual(len(data), 1)
            self.assertEqual(data[0]["Group"], "Group")
        finally:
            os.unlink(path)

    def test_special_characters_in_password(self):
        rows = [
            ["Title", "Username", "Password", "URL"],
            ["Server", "admin", "P@$$w0rd!#$%", "192.168.1.1"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["Password"], "P@$$w0rd!#$%")

    def test_empty_password(self):
        rows = [
            ["Title", "Username", "Password", "URL"],
            ["Server", "admin", "", "192.168.1.1"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["Password"], "")

    def test_ip_address_url_normalization(self):
        rows = [
            ["Title", "URL"],
            ["Server", "10.0.0.1"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["URL"], "http://10.0.0.1")

    def test_https_url_preserved(self):
        rows = [
            ["Title", "URL"],
            ["Server", "https://secure.example.com"],
        ]
        data = self._parse(rows, use_ai=False)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["URL"], "https://secure.example.com")


if __name__ == "__main__":
    unittest.main()
